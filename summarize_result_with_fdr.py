import matplotlib
import matplotlib.pyplot as plt
import os, fnmatch
import numpy as np
import pandas as pd
from statsmodels.stats.multitest import fdrcorrection
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook


SMALL_SIZE = 8
MEDIUM_SIZE = 10
BIGGER_SIZE = 18
sign_pair = ['<', '>=']
inequality_operators = {'<': lambda x, y: x < y,
                        '<=': lambda x, y: x <= y,
                        '>': lambda x, y: x > y,
                        '>=': lambda x, y: x >= y}

# plt.rc('xtick', labelsize=BIGGER_SIZE)
# plt.rc('ytick', labelsize=BIGGER_SIZE)
# plt.rc('legend', labelsize=BIGGER_SIZE)
plt.rcParams.update({'font.size': BIGGER_SIZE})


def find(pattern, path):
    result = []
    for root, dirs, files in os.walk(path):
        for name in files:
            if fnmatch.fnmatch(name, pattern):
                result.append(os.path.join(root, name))
    return result


def plot_histogram(asthma_df, plot_dir, pollutant_name, label_plot=['with asthma', 'w/o asthma']):
    # pollutant_col_name = asthma_df.columns
    df = asthma_df[[pollutant_name, 'label']]

    # pollutant_level_w_asthma = np.log10(df[asthma_df['label'] == 1][pollutant_name])
    # pollutant_level_wo_asthma = np.log10(df[asthma_df['label'] == 0][pollutant_name])
    pollutant_level_w_asthma = (df[asthma_df['label'] == 1][pollutant_name].values)
    pollutant_level_wo_asthma = (df[asthma_df['label'] == 0][pollutant_name].values)

    fig = plt.figure()
    ax = fig.add_subplot(111)
    _, bins, _ = ax.hist([pollutant_level_w_asthma, pollutant_level_wo_asthma],
                         weights=[np.ones(len(pollutant_level_w_asthma)) / len(pollutant_level_w_asthma),
                                  np.ones(len(pollutant_level_wo_asthma)) / len(pollutant_level_wo_asthma)],
                         label=label_plot,
                         color=['red', 'green'])

    ax.legend(loc='upper right')
    # ax.axvline(float(thres), color='k', linestyle='dashed', linewidth=1)
    min_ylim, max_ylim = ax.get_ylim()
    min_xlim, max_xlim = ax.get_xlim()
    y_pos = 0.75
    # ax.text(float(thres)+(max_xlim-min_xlim)*0.02, (max_ylim*y_pos+(1-y_pos)*min_ylim), 'Threshold = {:.3e}'.format(thres))
    # ax.set_title(pollutant_name+'_imputed_data')

    # pollutant_list = pollutant_name.split('y')
    # ax.yaxis.set_major_formatter(PercentFormatter(1))

    # ax.set_xlabel('{} level ({} scale)'.format(subscript_like_chemical_name(pollutant_list[0]), '${log_{10}}$'))
    ax.set_xlabel('{} level'.format(pollutant_name))
    ax.set_ylabel('Fraction of children exposed (separated by class)')
    # y_ticks = ax.get_yticks
    fig.savefig(os.path.join(plot_dir, "histogram_{}.png".format(pollutant_name)), bbox_inches="tight")


def heatmap(data, row_labels, col_labels, ax=None,
            cbar_kw={}, cbarlabel="", **kwargs):
    """
    Create a heatmap from a numpy array and two lists of labels.

    Parameters
    ----------
    data
        A 2D numpy array of shape (N, M).
    row_labels
        A list or array of length N with the labels for the rows.
    col_labels
        A list or array of length M with the labels for the columns.
    ax
        A `matplotlib.axes.Axes` instance to which the heatmap is plotted.  If
        not provided, use current axes or create a new one.  Optional.
    cbar_kw
        A dictionary with arguments to `matplotlib.Figure.colorbar`.  Optional.
    cbarlabel
        The label for the colorbar.  Optional.
    **kwargs
        All other arguments are forwarded to `imshow`.
    """

    if not ax:
        ax = plt.gca()

    # Plot the heatmap
    im = ax.imshow(data, **kwargs)

    # Create colorbar
    cbar = ax.figure.colorbar(im, ax=ax, spacing='proportional', **cbar_kw)
    cbar.ax.set_ylabel(cbarlabel, rotation=-90, va="bottom", fontsize=22)
    cbar.ax.tick_params(labelsize=22)

    # We want to show all ticks...
    ax.set_xticks(np.arange(data.shape[1]))
    ax.set_yticks(np.arange(data.shape[0]))
    # ... and label them with the respective list entries.
    ax.set_xticklabels(col_labels)
    ax.set_yticklabels(row_labels)

    # Let the horizontal axes labeling appear on top.
    ax.tick_params(top=True, bottom=False,
                   labeltop=True, labelbottom=False)

    # Rotate the tick labels and set their alignment.
    plt.setp(ax.get_xticklabels(), rotation=0,
             # ha="right",
             rotation_mode="anchor")

    # Turn spines off and create white grid.
    for edge, spine in ax.spines.items():
        spine.set_visible(False)
    ax.set_aspect(aspect='auto', adjustable='box')
    # ax.set_aspect(aspect='auto')
    # ax.set_aspect(aspect=0.3)
    # ax.set_xticks(np.arange(data.shape[1]+1)-.5, minor=True)
    # ax.set_yticks(np.arange(data.shape[0]+1)-.5, minor=True)
    ax.grid(which="minor", color="b", linestyle='-', linewidth=1)
    # ax.tick_params(which="minor", bottom=False, left=False)

    return im, cbar


def annotate_heatmap(im, data=None, valfmt="{x:.2f}",
                     textcolors=["black", "white"],
                     threshold=None, **textkw):
    """
    A function to annotate a heatmap.

    Parameters
    ----------
    im
        The AxesImage to be labeled.
    data
        Data used to annotate.  If None, the image's data is used.  Optional.
    valfmt
        The format of the annotations inside the heatmap.  This should either
        use the string format method, e.g. "$ {x:.2f}", or be a
        `matplotlib.ticker.Formatter`.  Optional.
    textcolors
        A list or array of two color specifications.  The first is used for
        values below a threshold, the second for those above.  Optional.
    threshold
        Value in data units according to which the colors from textcolors are
        applied.  If None (the default) uses the middle of the colormap as
        separation.  Optional.
    **kwargs
        All other arguments are forwarded to each call to `text` used to create
        the text labels.
    """

    if not isinstance(data, (list, np.ndarray)):
        data = im.get_array()

    # Normalize the threshold to the images color range.
    if threshold is not None:
        threshold = im.norm(threshold)
    else:
        threshold = im.norm(data.max()) / 2.

    # Set default alignment to center, but allow it to be
    # overwritten by textkw.
    kw = dict(horizontalalignment="center",
              verticalalignment="center",
              font_size=28)
    kw.update(textkw)

    # Get the formatter in case a string is supplied
    if isinstance(valfmt, str):
        valfmt = matplotlib.ticker.StrMethodFormatter(valfmt)

    # Loop over the data and create a `Text` for each "pixel".
    # Change the text's color depending on the data.
    texts = []
    for i in range(data.shape[0]):
        for j in range(data.shape[1]):
            kw.update(color=textcolors[int(im.norm(data[i, j]) > threshold)])
    #
    # text = im.axes.text(j, i, valfmt(data[i, j], None), **kw)
    #         texts.append(text)

    return texts


# fdr_path = './plot/nata_diagnose_year/fdr.csv'
# fdr_path = './plot/nata_birth_year/fdr.csv'


# outcome_list = ['act_score',
#                 'age_greaterthan5_diagnosed_asthma',
#                 # 'age_di'
#                 'daily_controller_past6months',
#                 'emergency_dept',
#                 'hospitalize_overnight',
#                 'regular_asthma_symptoms_past6months']
outcome_list = [
    # 'asthma',
    # 'asthma(act_score)',
    # 'age_greaterthan5_diagnosed_asthma',
    # 'age_diagnosed_asthma',
    'daily_controller_past6months',
    'emergency_dept',
    # 'emergency_dept_pastyr_count',
    'hospitalize_overnight',
    # 'hospitalize_overnight_pastyr_count',
    # 'regular_asthma_symptoms_past6months',
    # 'regular_asthma_symptoms_daysCount_pastWeek'
    # 'regular_asthma_symptoms_daysCount_pastWeek(greaterthan_nz_median)'
]

outcome_col_rename = {
    'asthma': 'asthma',
    # 'asthma(act_score)': 'act_score_(Cont.)',
    'age_greaterthan5_diagnosed_asthma': 'age_diagnosed_asthma_(>5 years old)',
    # 'age_diagnosed_asthma': 'age_diagnosed_asthma_(Cont.)',
    'daily_controller_past6months': 'daily_controller_past 6 months',
    'emergency_dept': 'emergency_dept_over_lifetime',
    # 'emergency_dept_pastyr_count': 'emergency_dept_pastyr_(Cont.)',
    'hospitalize_overnight': 'hospitalize_overnight_over_lifetime',
    # 'hospitalize_overnight_pastyr_count': 'hospitalize_overnight_pastyr_(Cont.)',
    # 'regular_asthma_symptoms_past6months': 'regular_asthma_symptoms_(Binary)',
    'regular_asthma_symptoms_daysCount_pastWeek(greaterthan_nz_median)': 'regular_asthma_symptoms_(past week)_(>non-zero median)'
    # 'regular_asthma_symptoms_daysCount_pastWeek': 'regular_asthma_symptoms_(Cont.)'
}

# p_val_cont = fdr_df.loc[fdr_df.apply(lambda x: '(Cont.)' in x['outcome'], axis=1), 'p_val'].values
# p_val_bin = fdr_df.loc[fdr_df.apply(lambda x: '(Binary)' in x['outcome'], axis=1), 'p_val'].values
#
# _, fdr_cont = fdrcorrection(p_val_cont)
# _, fdr_bin = fdrcorrection(p_val_bin)
#
# fdr_df.loc[fdr_df.apply(lambda x: '(Cont.)' in x['outcome'], axis=1), 'fdr'] = fdr_cont
# fdr_df.loc[fdr_df.apply(lambda x: '(Binary)' in x['outcome'], axis=1), 'fdr'] = fdr_bin
relation_dict = {'pos_correlate': 2.0,
                 # 'mixed_sign_profile': 0.0,
                 'neg_correlate': 1.0,
                 # 'na': 0
                 }


# relation_dict = {'Positive Coef.': 2.0,
#                  # 'mixed_sign_profile': 0.0,
#                  'Negative Coef.': 1.0,
#                  # 'na': 0
#                  }

# relation_dict_inv = {v:k for k, v in relation_dict.items()}
relation_list = [k for k in relation_dict]
# relation_list = ['NA'] + relation_list
# relation_list.append('NA')
relation_list.append('NA')

relation_inv_dict = {v: k for k, v in relation_dict.items()}


def summarize_plot(col='fdr', pollutant_suffix='', method_suffix='', outcome_col_table={}):
    print(pollutant_suffix)
    # for idx, outcome in enumerate(outcome_list):
    plot_dir = "./plot/{}/".format(method_suffix)
    pred_df = pd.read_csv('./pred_score_{}.csv'.format(method_suffix))
    outcome_rename_dummy = {}
    for k, v in outcome_col_table.items():
        outcome_rename_dummy[k] = '{}_{}{}'.format(v, '# of patients=',
                                                   pred_df.loc[pred_df['outcome'] == k, 'num_patients'].values[0])

    outcome_col_rename = outcome_rename_dummy
    # outcome_col_rename = {'act_score':'ACT',
    #                 'age_greaterthan5_diagnosed_asthma':'age>5',
    #                 'daily_controller_past6months': 'daily_controller',
    #                 'emergency_dept':'emerg_dept',
    #                 'hospitalize_overnight':'hos_night',
    #                 'regular_asthma_symptoms_past6months':'reg_symptoms'}

    outcome_reverse_dict = {v: k for k, v in outcome_col_rename.items()}

    # pollutant_list = ['EC', 'OC', 'SO4', 'NH4', 'Nit', 'NO2', 'PM2.5']

    fdr_path = './fdr_{}_count10.csv'.format(method_suffix)
    fdr_df = pd.read_csv(fdr_path, sep=',')
    fdr_df = fdr_df.loc[fdr_df['outcome'].isin(outcome_list)]
    print(fdr_df['outcome'].unique())
    fdr_df.loc[:,'fdr'] = fdrcorrection(fdr_df.loc[:,'p_val'].values)[-1]
    # print(fdr_df)
    # str_matched = fdr_df['profile'].str.contains(pollutant_suffix, regex=False)
    # print(str_matched)
    # fdr_df = fdr_df[str_matched]

    fdr_df = fdr_df.loc[fdr_df['fdr'] < 0.05]

    fdr_df = fdr_df.loc[fdr_df['outcome'] != 'asthma']
    fdr_df['profile'] = fdr_df['profile'].str.replace(pollutant_suffix, '', regex=False)
    single_pollutantTimeWin = dict()
    multi_pollutantTimeWin = dict()
    single_pollutant_fdr = dict()
    multi_pollutant_fdr = dict()
    single_pollutant_coef = dict()
    multi_pollutant_coef = dict()

    single_pollutant_freq = dict()
    multi_pollutant_freq = dict()
    for old_name, new_name in outcome_col_rename.items():
        fdr_df.replace(old_name, new_name, inplace=True)
    # print(fdr_df)
    new_outcome_list = [outcome_col_rename[old_name] for old_name in outcome_list]
    outcome_idx_dict = {out: idx for idx, out in enumerate(new_outcome_list)}

    list_profile_thres = dict()

    for row_idx, pollutant_row in fdr_df.iterrows():
        idx = outcome_idx_dict[pollutant_row['outcome']]
        pollutant_profile = pollutant_row['profile'].split('\t\t')
        # print(pollutant_profile, len(pollutant_profile))

        relation_to_outcome = pollutant_row['relation']
        # print(pollutant_row)
        if len(pollutant_profile) == 1:
            # set_p_fdr = pollutant_profile[0]
            set_p_fdr = pollutant_profile[0].split(sign_pair[1])[0].split(sign_pair[0])[0]
            print(set_p_fdr)
            # list_profile_thres[]
            if set_p_fdr in single_pollutant_fdr:
                single_pollutant_fdr[set_p_fdr][idx] = pollutant_row[col]
                single_pollutant_coef[set_p_fdr][idx] = abs(pollutant_row['coef'])
                single_pollutant_freq[set_p_fdr][idx] = abs(pollutant_row['freq'])
            else:
                single_pollutant_fdr[set_p_fdr] = np.ones(len(outcome_list))
                single_pollutant_fdr[set_p_fdr][idx] = pollutant_row[col]
                single_pollutant_coef[set_p_fdr] = np.zeros(len(outcome_list))
                single_pollutant_coef[set_p_fdr][idx] = abs(pollutant_row['coef'])
                single_pollutant_freq[set_p_fdr] = np.zeros(len(outcome_list))
                single_pollutant_freq[set_p_fdr][idx] = abs(pollutant_row['freq'])
        else:
            # print(pollutant_profile)
            p_thres = []
            p_wo_thres = []
            # set_p_with_sign = []
            for p in pollutant_profile:
                if sign_pair[1] in p:
                    p_splitted_by_sign = p.split(sign_pair[1])
                    p_wo_thres.append(p_splitted_by_sign[0] + sign_pair[1])
                else:
                    p_splitted_by_sign = p.split(sign_pair[0])
                    p_wo_thres.append(p_splitted_by_sign[0] + sign_pair[0])
                p_thres.append(float(p_splitted_by_sign[1]))

            # set_p_fdr = frozenset(p_wo_thres)
            set_p_fdr = tuple(p_wo_thres)
            # set_p_fdr = tuple(pollutant_profile)
            if set_p_fdr in multi_pollutant_fdr:
                multi_pollutant_fdr[set_p_fdr][idx] = pollutant_row[col]
                multi_pollutant_coef[set_p_fdr][idx] = abs(pollutant_row['coef'])
                multi_pollutant_freq[set_p_fdr][idx] = abs(pollutant_row['freq'])
            else:
                multi_pollutant_fdr[set_p_fdr] = np.ones(len(outcome_list))
                multi_pollutant_fdr[set_p_fdr][idx] = pollutant_row[col]
                multi_pollutant_coef[set_p_fdr] = np.zeros(len(outcome_list))
                multi_pollutant_coef[set_p_fdr][idx] = abs(pollutant_row['coef'])
                multi_pollutant_freq[set_p_fdr] = np.zeros(len(outcome_list))
                multi_pollutant_freq[set_p_fdr][idx] = abs(pollutant_row['freq'])



            # if pollutant_row['coef'] > 0:
            #     rel = 2.0
            # else:
            #     rel = 1.0
            # if pollutant_row[col] > 0.05:
            #     rel = 0.5

            # rel = relation_dict[relation_to_outcome]
            # if

        if len(pollutant_profile) == 1:
            set_p = pollutant_profile[0].split(sign_pair[1])[0].split(sign_pair[0])[0]
            if pollutant_row[col] < 0.05:
                # print(set_p)
                if relation_to_outcome == 'mixed_sign_profile':
                    rel = 0.5
                else:
                    rel = relation_dict[relation_to_outcome]
                    # print(set_p)

            if sign_pair[1] in pollutant_profile[0]:
                p_splitted_by_sign = pollutant_profile[0].split(sign_pair[1])
                p_wo_thres = p_splitted_by_sign[0] + sign_pair[1]
            else:
                p_splitted_by_sign = pollutant_profile[0].split(sign_pair[0])
                p_wo_thres = p_splitted_by_sign[0] + sign_pair[0]
            p_thres = float(p_splitted_by_sign[1])
            # set_p = pollutant_profile[0]
            if set_p in single_pollutantTimeWin:
                single_pollutantTimeWin[set_p][idx] = rel
            else:
                single_pollutantTimeWin[set_p] = 0.5 * np.ones(len(outcome_list))
                single_pollutantTimeWin[set_p][idx] = rel
        else:
            if pollutant_row['coef'] > 0:
                rel = 2.0
            else:
                rel = 1.0
            if pollutant_row[col] > 0.05:
                rel = 0.5

            p_wo_thres = []
            # print(pollutant_profile)
            for p in pollutant_profile:
                if sign_pair[1] in p:
                    p_splitted_by_sign = p.split(sign_pair[1])
                    p_wo_thres.append(p_splitted_by_sign[0] + sign_pair[1])
                else:
                    p_splitted_by_sign = p.split(sign_pair[0])
                    p_wo_thres.append(p_splitted_by_sign[0] + sign_pair[0])
                p_thres.append(float(p_splitted_by_sign[1]))
            p_wo_thres = tuple(p_wo_thres)
            # for p in pollutant_profile:
            #     if sign_pair[1] in p:
            #         p_wo_thres.append(p.split(sign_pair[1])[0]+sign_pair[1])
            #     else:
            #         p_wo_thres.append(p.split(sign_pair[0])[0] + sign_pair[0])
            set_p = tuple(p_wo_thres)
            # set_p = frozenset(p_wo_thres)
            # set_p = tuple(pollutant_profile)
            if set_p in multi_pollutantTimeWin:
                multi_pollutantTimeWin[set_p][idx] = rel
            else:
                multi_pollutantTimeWin[set_p] = 0.5 * np.ones(len(outcome_list))
                multi_pollutantTimeWin[set_p][idx] = rel

        # if p_wo_thres in list_profile_thres:
        # TODO: update to be capable for same profiles and threshold
        list_profile_thres[p_wo_thres] = (p_thres, outcome_list[idx], pollutant_row['coef'])
                # list_profile_thres[p_wo_thres].append((p_thres, outcome_list[idx], pollutant_row['coef']))
            # else:
            #     list_profile_thres[p_wo_thres] = [(p_thres, outcome_list[idx], pollutant_row['coef'])]

    m_keys = multi_pollutantTimeWin.copy().keys()
    # print(m_keys)
    s_keys = single_pollutantTimeWin.copy().keys()
    # print(s_keys)
    #
    # set_of_pollutants = set(list(m_keys))
    # set_of_pollutants.update(s_keys)
    #
    # set_of_pollutants_outcomes = {}
    # for s in s_keys:
    #     for o_idx, outcome in enumerate(outcome_list):
    #         if single_pollutantTimeWin[s][o_idx] != 0.5:
    #             # set_of_pollutants_outcomes[s] = (outcome, single_pollutantTimeWin[s][o_idx])
    #             set_of_pollutants_outcomes[s] = (outcome, single_pollutant_coef[s][o_idx])
    #
    # for m in m_keys:
    #     for o_idx, outcome in enumerate(outcome_list):
    #         if multi_pollutantTimeWin[m][o_idx] != 0.5:
    #             # set_of_pollutants_outcomes[m] = (outcome, multi_pollutantTimeWin[m][o_idx])
    #             set_of_pollutants_outcomes[m] = (outcome, multi_pollutant_coef[m][o_idx])

    if len(m_keys) > 0:
        for key in m_keys:
            new_key = ''
            k_list = list(key)
            # k_list.sort()
            for k in k_list:
                new_key += '{}\n'.format(k)
            new_key = new_key[:-1]
            multi_pollutantTimeWin[new_key] = multi_pollutantTimeWin.pop(key)

    m_keys = multi_pollutant_fdr.copy().keys()
    if len(m_keys) > 0:
        for key in m_keys:
            new_key = ''
            k_list = list(key)
            # k_list.sort()
            for k in k_list:
                new_key += '{}\n'.format(k)
            new_key = new_key[:-1]
            multi_pollutant_fdr[new_key] = multi_pollutant_fdr.pop(key)

    m_keys = multi_pollutant_coef.copy().keys()
    if len(m_keys) > 0:
        for key in m_keys:
            new_key = ''
            k_list = list(key)
            # k_list.sort()
            for k in k_list:
                new_key += '{}\n'.format(k)
            new_key = new_key[:-1]
            multi_pollutant_coef[new_key] = multi_pollutant_coef.pop(key)

    m_keys = multi_pollutant_freq.copy().keys()
    if len(m_keys) > 0:
        for key in m_keys:
            new_key = ''
            k_list = list(key)
            # k_list.sort()
            for k in k_list:
                new_key += '{}\n'.format(k)
            new_key = new_key[:-1]
            multi_pollutant_freq[new_key] = multi_pollutant_freq.pop(key)

    # single_df = pd.DataFrame.from_dict(single_pollutantTimeWin,
    #                                    orient='index',
    #                                    columns=new_outcome_list)
    # multi_df = pd.DataFrame.from_dict(multi_pollutantTimeWin,
    #                                   orient='index',
    #                                   columns=new_outcome_list)
    #
    #
    # single_fdr_df = pd.DataFrame.from_dict(single_pollutant_fdr,
    #                                    orient='index',
    #                                    columns=new_outcome_list)
    # multi_fdr_df = pd.DataFrame.from_dict(multi_pollutant_fdr,
    #                                   orient='index',
    #                                   columns=new_outcome_list)
    #
    # single_coef_df = pd.DataFrame.from_dict(single_pollutant_coef,
    #                                        orient='index',
    #                                        columns=new_outcome_list)
    # multi_coef_df = pd.DataFrame.from_dict(multi_pollutant_coef,
    #                                       orient='index',
    #                                       columns=new_outcome_list)

    # def sorted_by_freq(hm_df, ):

    single_df = pd.DataFrame.from_dict(single_pollutantTimeWin,
                                       orient='index',
                                       columns=new_outcome_list).sort_index()
    multi_df = pd.DataFrame.from_dict(multi_pollutantTimeWin,
                                      orient='index',
                                      columns=new_outcome_list).sort_index()
    single_fdr_df = pd.DataFrame.from_dict(single_pollutant_fdr,
                                           orient='index',
                                           columns=new_outcome_list).sort_index()
    multi_fdr_df = pd.DataFrame.from_dict(multi_pollutant_fdr,
                                          orient='index',
                                          columns=new_outcome_list).sort_index()

    single_coef_df = pd.DataFrame.from_dict(single_pollutant_coef,
                                            orient='index',
                                            columns=new_outcome_list).sort_index()
    multi_coef_df = pd.DataFrame.from_dict(multi_pollutant_coef,
                                           orient='index',
                                           columns=new_outcome_list).sort_index()

    single_freq_df = pd.DataFrame.from_dict(single_pollutant_freq,
                                            orient='index',
                                            columns=new_outcome_list).sort_index()
    multi_freq_df = pd.DataFrame.from_dict(multi_pollutant_freq,
                                           orient='index',
                                           columns=new_outcome_list).sort_index()

    # print(single_df)
    # print(single_df)
    # print(multi_df)
    # merged_df = pd.concat([single_df, multi_df])
    # merged_fdr_df = pd.concat([single_fdr_df, multi_fdr_df])
    # merged_coef_df = pd.concat([single_coef_df, multi_coef_df])
    # merged_freq_df = pd.concat([single_freq_df, multi_freq_df])

    merged_df = pd.concat([multi_df, single_df])
    merged_fdr_df = pd.concat([multi_fdr_df, single_fdr_df])
    merged_coef_df = pd.concat([multi_coef_df, single_coef_df])
    merged_freq_df = pd.concat([multi_freq_df, single_freq_df])

    # merged_fdr_df = merged_fdr_df.str.cat(merged_freq_df.astype(str), sep='\nCount=')
    # print(fdr_df)
    # print(fdr_df.loc[:, 'max_count'])
    max_count = fdr_df.loc[:, 'max_count'].values[0]
    #
    # merged_freq_df['max'] = merged_freq_df.max(axis=1).astype(int).astype(str)
    # # merged_freq_df.sort_values(by='max', ascending=False, inplace=True)
    #
    # new_idx_name = merged_freq_df.index.str.cat(merged_freq_df[['max']],
    #                                             sep='\n(Frequency = ')
    # new_idx_name = new_idx_name + ' out of {})'.format(max_count)

    # merged_df = merged_df.loc[merged_freq_df.index]
    # merged_df.index = new_idx_name
    #
    # # merged_coef_df = merged_coef_df.loc[merged_freq_df.index]
    # merged_coef_df.index = new_idx_name
    #
    # # merged_fdr_df = merged_fdr_df.loc[merged_freq_df.index]
    # merged_fdr_df.index = new_idx_name
    # merged_freq_df.index = new_idx_name

    # merged_df.rename(columns=outcome_col_rename, inplace=True)

    fontsize_pt = plt.rcParams['xtick.labelsize']
    dpi = 72.27

    # comput the matrix height in points and inches
    # matrix_height_pt = fontsize_pt * merged_df.shape[1]
    # print(matrix_height_pt)
    # matrix_height_in = matrix_height_pt / dpi

    # compute the required figure height
    top_margin = 0.04  # in percentage of the figure height
    bottom_margin = 0.04  # in percentage of the figure height

    # figure_height = matrix_height_in / (1 - top_margin - bottom_margin)
    # plt.clf()
    def fig_gen(merged_temp_df):
        fig = plt.figure(figsize=(32, 40))
        ax = fig.add_subplot(111)
        qrates = np.array(relation_list)
        np_array = [0.5 - 1e-4, 0.5 + 1e-4, 1.5, 2.5]
        norm = matplotlib.colors.BoundaryNorm(np_array, 3)
        fmt = matplotlib.ticker.FuncFormatter(lambda x, pos: qrates[::-1][norm(x)])

        im, _ = heatmap(merged_temp_df.values, merged_temp_df.index, [c.replace('_', '\n') for c in merged_temp_df.columns], ax=ax,
                        cmap=matplotlib.colors.ListedColormap(['white', 'red', 'green']),
                        norm=norm,
                        cbar_kw=dict(ticks=np.arange(1, 3), format=fmt, ),
                        cbarlabel="Association of Profile to the Outcome")
        ax.tick_params(axis='x', labelsize=25)
        ax.tick_params(axis='y', labelsize=25)

        return fig, ax, im

    # output file
    max_rows_in1 = 20

    number_profiles = merged_df.shape[0]
    idx_range = np.append(np.arange(0, number_profiles, max_rows_in1), number_profiles)
    number_of_outcomes_by_profile = (merged_df != 0.5).astype(int).sum(axis=1)
    number_of_figures = np.ceil(float(number_profiles) / max_rows_in1).astype(int)

    merged_df['num_outcomes'] = number_of_outcomes_by_profile
    # print(number_of_outcomes_by_profile)

    profile_str = merged_df.index.str
    merged_df['all_greater'] = (profile_str.count('>') > 1) & (profile_str.count('<') == 0)
    merged_df['all_less'] = (profile_str.count('<') > 1) & (profile_str.count('>') == 0)
    # merged_df['multi_pollutants'] = False
    merged_df['multi_pollutants'] = profile_str.count('\n') > 0
    # merged_df['multi_pollutants'] = (profile_str.count('>') < 1) | (profile_str.count('<') < 1)
    sorted_by_columns = ['all_greater', 'all_less','multi_pollutants', 'num_outcomes']
    merged_df.sort_values(by=sorted_by_columns,
                                             ascending=[False, False, False, False],
                                             inplace=True)

    merged_fdr_df = merged_fdr_df.loc[merged_df.index]
    merged_coef_df = merged_coef_df.loc[merged_df.index]
    merged_freq_df = merged_freq_df.loc[merged_df.index]

    # merged_df.drop(columns=sorted_by_columns, inplace=True)
    profile_str_sorted = merged_df.index.str
    category_outcome = {'all_greater': merged_df['all_greater'],
                        'all_less': merged_df['all_less'],
                        'mixed_sign_multi_pollutants': (merged_df['multi_pollutants'] & ~merged_df['all_greater'] & ~merged_df['all_less']),
                        'single_pollutant': ~merged_df['multi_pollutants']}
    print(sum(~merged_df['multi_pollutants']))
    merged_df.drop(columns=sorted_by_columns, inplace=True)
    if not os.path.exists(plot_dir):
        os.mkdir(plot_dir)
    # writer = pd.ExcelWriter('fdr_summary.xlsx', engine='xlsxwriter')

    wb = Workbook()


    for profile_cat, profile_bool in category_outcome.items():
        cat_dir = os.path.join(plot_dir, profile_cat)
        if not os.path.exists(cat_dir):
            os.mkdir(cat_dir)
        merged_df_cat = merged_df[profile_bool]
        merged_fdr_df_cat = merged_fdr_df[profile_bool]
        merged_coef_df_cat = merged_coef_df[profile_bool]
        merged_freq_df_cat = merged_freq_df[profile_bool]

        number_profiles = merged_df_cat.shape[0]
        idx_range = np.append(np.arange(0, number_profiles, max_rows_in1), number_profiles)
        number_of_figures = np.ceil(float(number_profiles) / max_rows_in1).astype(int)

        for fig_sub_idx in range(number_of_figures):
            row_start_idx = idx_range[fig_sub_idx]
            row_end_idx = idx_range[fig_sub_idx+1]
            merged_df_sub = merged_df_cat.iloc[row_start_idx:row_end_idx].copy()
            merged_fdr_df_sub = merged_fdr_df_cat.iloc[row_start_idx:row_end_idx].copy()
            merged_coef_df_sub = merged_coef_df_cat.iloc[row_start_idx:row_end_idx].copy()
            merged_freq_df_sub = merged_freq_df_cat.iloc[row_start_idx:row_end_idx].copy()
            # print(merged_df_sub)

            fig1, ax1, im1 = fig_gen(merged_temp_df=merged_df_sub)
            fig2, ax2, im2 = fig_gen(merged_temp_df=merged_df_sub)
            fig1.suptitle('{} Significant Pollutant Profile ({} < 0.05)'.format(pollutant_suffix, col), fontsize=20, y=.95)
            fig2.suptitle('Coefficients of {} Significant Pollutant Profile'.format(pollutant_suffix), fontsize=20, y=.95)
            # print(merged_coef_df)
            # convert_str = {c: str for c in merged_df.columns}
            # merged_df = merged_df.astype(convert_str)
            # for k, v in relation_dict.items():
            #     merged_df_sub = merged_df_sub.replace(str(v), k)
            # merged_df_sub.replace('0.5', '>0.05', inplace=True)
            # yr_dir = pollutant_suffix.replace('+5', '_yr_5').replace('-5', '_yr_-5')
            # plot_dir = './plot/nata_{}/histogram_profile_from_{}/'.format(method_suffix, yr_dir, pollutant_suffix)
            # if 'birth' in pollutant_suffix:
            #     plot_dir = plot_dir.format('birth')
            # else:
            #     plot_dir = plot_dir.format('diagnose')



            for idx_r, r in enumerate(merged_df_sub.index):
                for idx_c, c in enumerate(merged_df_sub.columns):
                    if merged_fdr_df_sub.loc[r, c] < 0.05:
                        plot_txt_fdr = 'FDR={:.2e}\nCount={:d}'.format(merged_fdr_df_sub.loc[r, c], int(merged_freq_df_sub.loc[r, c]))
                        plot_txt_coef = '{:.3f}\nCount={:d}'.format(merged_coef_df_sub.loc[r, c], int(merged_freq_df_sub.loc[r, c]))
                        im1.axes.text(idx_c, idx_r, plot_txt_fdr, ha="center", va="center", color="w", size=28,
                                      fontweight='bold')
                        im2.axes.text(idx_c, idx_r, plot_txt_coef, ha="center", va="center", color="w", size=28,
                                      fontweight='bold')
                # if '\n' in r:
                r_list = r.split('\n')
                for sub_r in r_list:
                    sub_r = sub_r.replace(sign_pair[0], '').replace(sign_pair[1], '')
                    # plot_histogram(asthma_df=asthma_df, plot_dir=plot_dir, pollutant_name=sub_r+'birth+5')

            plt.autoscale()
            fig1.savefig(os.path.join(cat_dir,'heatmp_summary_nata_{}_{}_{}.pdf'.format(pollutant_suffix, col, fig_sub_idx)),
                         bbox_inches="tight")
            # fig2.savefig(os.path.join(cat_dir,'heatmp_summary_nata_{}_{}_{}.pdf'.format(pollutant_suffix, 'coef', fig_sub_idx)),
            #              bbox_inches="tight")
        ws_fdr = wb.create_sheet(profile_cat+'_fdr')
        merged_fdr_df_cat_replaced = merged_fdr_df_cat.replace(1, np.nan)
        for r in dataframe_to_rows(merged_fdr_df_cat_replaced, index=True, header=True):
            ws_fdr.append(r)
        merged_rel_df_cat_replaced = merged_df_cat.replace({2.0:'Positively Associated',
                                                            1.0:'Negatively Associated',
                                                            0.5:'NA'})
        ws_asso = wb.create_sheet(profile_cat + '_association')
        for r in dataframe_to_rows(merged_rel_df_cat_replaced, index=True, header=True):
            ws_asso.append(r)

        merged_freq_df_cat_replaced = merged_freq_df_cat.replace(0, np.nan)
        ws_freq = wb.create_sheet(profile_cat + '_frequency')
        for r in dataframe_to_rows(merged_freq_df_cat_replaced, index=True, header=True):
            ws_freq.append(r)

        # merged_fdr_df_cat.to_csv(os.path.join(cat_dir,'summary_{}_{}.csv'.format(col, pollutant_suffix)))

    wb.save(os.path.join(plot_dir, "summary.xlsx"))
    ############ new plot by outcome ###############
    # max_rows = 10
    # max_col_sub_fig = 3
    # print(merged_df)
    # for idx_c, c in enumerate(merged_df.columns):
    #     # print(c)
    #     merged_fdr_df_c = merged_fdr_df[str(c)]
    #     print(merged_fdr_df_c)
    #     merge_df_significant_idx = merged_fdr_df_c < 0.05
    #
    #     merged_fdr_df_c_significant = merged_fdr_df_c.loc[merge_df_significant_idx]
    #     merged_df_c_significant = merged_df.loc[merge_df_significant_idx, [c]]
    #
    #     ### Sorted by FDR, all greater, Freq
    #     merged_freq_df_c_significant = merged_freq_df.loc[merge_df_significant_idx, :]
    #     # merged_freq_df_c_significant.sort_values(ascending=False, inplace=True)
    #
    #     # merged_freq_df_c_significant['freq'] = merged_freq_df_c_significant.max(axis=1).astype(int).astype(str)
    #     # merged_freq_df.sort_values(by='max', ascending=False, inplace=True)
    #     # print()
    #     new_idx_name = merged_freq_df_c_significant.index.str.cat(merged_freq_df[[c]].astype(str),
    #                                                 sep='\n(Frequency = ')
    #     new_idx_name = new_idx_name + ' out of {})'.format(max_count)
    #
    #     merged_freq_df_c_significant.index = new_idx_name
    #     merged_fdr_df_c_significant.index = new_idx_name
    #     merged_df_c_significant.index = new_idx_name
    #
    #     profile_str = merged_freq_df_c_significant.index.str
    #     merged_freq_df_c_significant['all_greater'] = False
    #     # print(profile_str.count('>') > 1)
    #     # print(profile_str.count('<') == 0)
    #     merged_freq_df_c_significant['all_greater'] = (profile_str.count('>') > 1) & (profile_str.count('<') == 0)
    #     merged_freq_df_c_significant['multi_pollutants'] = False
    #     merged_freq_df_c_significant['multi_pollutants'] = profile_str.count('\n') > 1
    #     merged_freq_df_c_significant.sort_values(by=['multi_pollutants','all_greater',  c],
    #                                              ascending=[False, False, False],
    #                                              inplace=True)
    #
    #     merged_fdr_df_c_significant = merged_fdr_df_c_significant.loc[merged_freq_df_c_significant.index]
    #     merged_df_c_significant = merged_df_c_significant.loc[merged_freq_df_c_significant.index]
    #
    #
    #     number_profiles_c = sum(merged_fdr_df_c < 0.05)
    #     print(number_profiles_c, merged_df_c_significant.shape, merged_fdr_df_c_significant.shape)
    #     # idx_range = np.linspace(0, number_profiles_c, num=max_rows, dtype=int)
    #     idx_range = np.append(np.arange(0, number_profiles_c, max_rows), number_profiles_c)
    #     num_sub_figs = len(idx_range)-1
    #     row_sub_fig = np.ceil(num_sub_figs / max_col_sub_fig).astype(int)
    #     col_sub_fig = min(max_col_sub_fig, num_sub_figs)
    #     fig_outcome, ax_outcome = plt.subplots(row_sub_fig, max_col_sub_fig, figsize=(17*col_sub_fig, 13*row_sub_fig))
    #     for idx_ra in range(num_sub_figs):
    #         subfig_bool = np.zeros((merged_df_c_significant.shape[0])).astype(bool)
    #         subfig_bool[idx_range[idx_ra]:idx_range[idx_ra+1]] = True
    #         merged_df_subfig = merged_df_c_significant.iloc[idx_range[idx_ra]:idx_range[idx_ra+1]]
    #         merged_df_fdr_subfig = merged_fdr_df_c_significant.iloc[idx_range[idx_ra]:idx_range[idx_ra+1]]
    #         # print(merged_df_c_significant)
    #         # print(merged_df_subfig)
    #         subfig_r = idx_ra // max_col_sub_fig
    #         subfig_c = idx_ra % max_col_sub_fig
    #         print(subfig_c, subfig_r)
    #         ax_subfig = ax_outcome[subfig_r, subfig_c]
    #         qrates = np.array(relation_list)
    #         np_array = [0.5 - 1e-4, 0.5 + 1e-4, 1.5, 2.5]
    #         norm = matplotlib.colors.BoundaryNorm(np_array, 3)
    #         fmt = matplotlib.ticker.FuncFormatter(lambda x, pos: qrates[::-1][norm(x)])
    #
    #         im_subfig, cbar_subfig = heatmap(np.expand_dims(merged_df_subfig.values,axis=1),
    #                                merged_df_subfig.index,
    #                         # [c.replace('_', '\n') for _col in merged_df_subfig.columns],
    #                         [''],
    #                         ax=ax_subfig,
    #                         cmap=matplotlib.colors.ListedColormap(['white', 'red', 'green']),
    #                         norm=norm,
    #                         cbar_kw=dict(ticks=np.arange(1, 3), format=fmt),
    #                         cbarlabel="Coefficient of Profile to the Outcome")
    #         # ax_subfig.tick_params(axis='x', labelsize=25)
    #         # ax_subfig.tick_params(axis='y', labelsize=25)
    #         if (idx_ra != (num_sub_figs-1)) and (subfig_c != (max_col_sub_fig-1)):
    #             cbar_subfig.remove()
    #
    #         for idx_r, r in enumerate(merged_df_fdr_subfig.index):
    #             # for idx_c, c in enumerate(merged_df_fdr_subfig.columns):
    #             # if merged_fdr_df.loc[r] < 0.05:
    #             plot_txt_fdr = '{:.3e}'.format(merged_df_fdr_subfig.loc[r])
    #             # plot_txt_coef = '{:.3f}'.format(merged_coef_df.loc[r, c])
    #             im_subfig.axes.text(0, idx_r, plot_txt_fdr, ha="center", va="center", color="w", size=18,
    #                           fontweight='bold')
    #                     # im2.axes.text(idx_c, idx_r, plot_txt_coef, ha="center", va="center", color="w", size=28,
    #                     #               fontweight='bold')
    #     empty_subfig = range(num_sub_figs, row_sub_fig*max_col_sub_fig)
    #     for i in empty_subfig:
    #         fig_outcome.delaxes(ax_outcome.flatten()[i])
    #     # fig_outcome.subplots_adjust()
    #     fig_outcome.suptitle('{} Significant Pollutant Profile of {} ({} < 0.05)'.format(pollutant_suffix, c, col), fontsize=25, y=0.94)
    #     fig_outcome.tight_layout(pad=5.0)
    #     fig_outcome.savefig("./plot/{}/summary_fig_{}_{}_{}.pdf".format(method_suffix, c, pollutant_suffix, col),bbox_inches="tight")




    return list_profile_thres


# sig_list = ['p_val', 'fdr']
#
# def search_outcomes_by_pollutant(df, pollutant_set, )

# method_suffix_list = ['bt_xgb_multiple_counts',
#                       'nbt_xgb_multiple_counts',
#                       'bt_xgb_single_count',
#                       'nbt_xgb_single_count']

method_suffix_list = ['bt_multiple',
                      'nbt_multiple',
                      'bt_single',
                      'nbt_single']

sig_list = ['fdr']
suffix_list = ['birth+5']
asthma_csv_path = 'data/asthma_NATA_birth_yr_5.csv'
asthma_df = pd.read_csv(asthma_csv_path)
set_pollutants_list = []
set_pollutants_with_outcome_list = []

list_of_profile_detail = []
for m in method_suffix_list:
    for sig in sig_list:
        for suffix in suffix_list:
            profile_thres = summarize_plot(sig, pollutant_suffix=suffix,
                                           method_suffix=m, outcome_col_table=outcome_col_rename)
            set_pollutants_list.append(set(profile_thres.keys()))
            list_of_profile_detail.append(profile_thres)
            # set_pollutants_with_outcome_list.append(set_pollutants_with_outcome)

intersect_pollutant_list = set.intersection(*set_pollutants_list)
print(intersect_pollutant_list)

intersect_pollutant_dictionary = {}
profile_output_str = 'data/{}_{}_profile_label.csv'
asthma_id_df = pd.read_csv('asthma_subject_ID.csv')
asthma_id_df['Study ID'] = asthma_id_df['Study ID'].astype(str)

for outcome in outcome_list:
    original_path = 'data/{}_NATA_birth_yr_5.csv'.format(outcome)
    orig_df = pd.read_csv(original_path).reset_index()
    orig_df['ID'] = orig_df['ID'].astype(int).astype(str)
    # orig_df['ID'] = orig_df['ID'].map('N{:03d}'.format)
    # orig_df['ID'] = orig_df['ID'].astype(int).map('N{:03d}'.format)
    # orig_df['ID_temp'] = ''
    for index, row in orig_df.iterrows():
        # print(row['ID'])
        # print(orig_df.loc[index, 'ID'])
        # print(asthma_id_df.loc[asthma_id_df['Study ID'] == row['ID']])
        # print(asthma_id_df.loc[asthma_id_df['Study ID'] == row['ID'],'Assigned Subject _ID (e.g. N001, N002, etc.)'].values)
        orig_df.loc[index, 'ID'] = asthma_id_df.loc[asthma_id_df['Study ID'] == row['ID'],
                                                    'Assigned Subject _ID (e.g. N001, N002, etc.)'].values

    for m_idx, m in enumerate(list_of_profile_detail):
        method_thres_df = orig_df[['ID', 'label']]
        for k, v in m.items():
            if v[1] == outcome:
                print(k, v)
                single_p_only = False
                if type(k) != str:
                    profile_str = '\n'.join(['{}{:.3e}'.format(single_p, v[0][idx]) for idx, single_p in enumerate(k)])
                else:
                    profile_str = '{}{:.3e}'.format(k, v[0])
                    single_p_only = True
                bool_profile = np.ones(method_thres_df.shape[0]).astype(int)
                if single_p_only:
                    if sign_pair[0] in k:
                        sign_notation = sign_pair[0]
                    else:
                        sign_notation = sign_pair[1]
                    bool_profile = inequality_operators[sign_notation](orig_df[k.replace(sign_notation, '')], v[0])
                else:
                    for idx, sp in enumerate(k):
                        if sign_pair[0] in sp:
                            sign_notation = sign_pair[0]
                        else:
                            sign_notation = sign_pair[1]
                        current_bool = inequality_operators[sign_notation](orig_df[sp.replace(sign_notation, '')],
                                                                           v[0][idx])
                        bool_profile = bool_profile & current_bool
                # method_thres_df[profile_str] = int(bool_profile)
                profile_str = profile_str.replace('[', '(')
                profile_str = profile_str.replace(']', ')')
                method_thres_df[profile_str] = bool_profile.astype(int)
        method_thres_df.to_csv(profile_output_str.format(method_suffix_list[m_idx], outcome), index=False)

    for i in intersect_pollutant_list:
        outcome_for_this_profile = list_of_profile_detail[0][i][1]
        if outcome == outcome_for_this_profile:
            bool_profile = np.ones(method_thres_df.shape[0]).astype(bool)
            intersect_thres_df = orig_df[['ID', 'label']]
            thres_mat = np.array([profile_detail[i][0] for profile_detail in list_of_profile_detail])
            coef_mat = np.array([profile_detail[i][2] for profile_detail in list_of_profile_detail])
            if np.median(coef_mat) > 0:
                coef_sign = 'Positive'
            else:
                coef_sign = 'Negative'
            if type(i) == tuple:
                thres_median = np.median(thres_mat, axis=0)
                key = '\n'.join(['{}{:.3e}'.format(single_p, thres_median[idx]) for idx, single_p in enumerate(i)])
                multi = True
                for idx, sp in enumerate(i):
                    if sign_pair[0] in sp:
                        sign_notation = sign_pair[0]
                    else:
                        sign_notation = sign_pair[1]
                    current_bool = inequality_operators[sign_notation](orig_df[sp.replace(sign_notation, '')],
                                                                       thres_median[idx])
                    bool_profile = bool_profile & current_bool
            else:
                key = '{}{:.3e}'.format(i, np.median(thres_mat))
                multi = False
                if sign_pair[0] in i:
                    sign_notation = sign_pair[0]
                else:
                    sign_notation = sign_pair[1]
                bool_profile = inequality_operators[sign_notation](orig_df[i.replace(sign_notation, '')],
                                                                   np.median(thres_mat))

            intersect_thres_df[key] = bool_profile
            intersect_thres_df.to_csv(profile_output_str.format('intersect', outcome), index=False)

for i in intersect_pollutant_list:
    # for m_idx, m in enumerate(method_suffix_list):
    # print('{}: {}'.format(i, set_pollutants_with_outcome_list[m_idx][i]))
    thres_mat = np.array([profile_detail[i][0] for profile_detail in list_of_profile_detail])
    coef_mat = np.array([profile_detail[i][2] for profile_detail in list_of_profile_detail])
    if np.median(coef_mat) > 0:
        coef_sign = 'Positive'
    else:
        coef_sign = 'Negative'
    if type(i) == tuple:
        thres_median = np.median(thres_mat, axis=0)
        key = '\n'.join(['{}{:.3e}'.format(single_p, thres_median[idx]) for idx, single_p in enumerate(i)])
        multi = True
    else:
        key = '{}{:.3e}'.format(i, np.median(thres_mat))
        multi = False
    # sp = set_pollutants_with_outcome_list[0][i]

    intersect_pollutant_dictionary[key] = (list_of_profile_detail[0][i][1], coef_sign, multi)

print(intersect_pollutant_dictionary)
intersect_df = pd.DataFrame({'mutually inclusive profile': [k for k, v in intersect_pollutant_dictionary.items()],
                             'outcome': [v[0] for k, v in intersect_pollutant_dictionary.items()],
                             'coef': [v[1] for k, v in intersect_pollutant_dictionary.items()],
                             'combination': [v[2] for k, v in intersect_pollutant_dictionary.items()],
                             }
                            )

category_outcome = pd.api.types.CategoricalDtype(categories=outcome_list, ordered=True)
intersect_df['outcome'] = intersect_df['outcome'].astype(category_outcome)
intersect_df.sort_values(by=['outcome'], inplace=True)
intersect_df.sort_values(by=['combination'], inplace=True, ascending=False)

# print(intersect_df)
intersect_df.to_csv('intersection_of_extracted_profile.csv', index=False)
# summarize_plot('p_val')
# summarize_plot('fdr')
